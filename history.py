import dash                                                   # pip3 install dash
import dash_core_components as dcc
import dash_html_components as html
import dash_table

from dash.dependencies import Output, State, Input

import plotly.graph_objects as go                                     
from plotly.subplots import make_subplots
import plotly.express as px

import pandas as pd
from datetime import datetime
import matplotlib.dates as mdates
import alpaca_trade_api as alpaca                                     # pip3 install alpaca-trade-api

from ipywidgets import interactive, HBox, VBox
import openpyxl as xl
from openpyxl import Workbook, load_workbook                                         # pip3 install openpyxl
import yaml

# for plotting transform plots
import numpy as np
import math
from scipy import signal
import matplotlib
matplotlib.use('TkAgg') # pip3 install tk
import matplotlib.pyplot as plt
import tkinter
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)
# Implement the default Matplotlib key bindings.
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure

external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']


app = dash.Dash(__name__, external_stylesheets=external_stylesheets, 
                    suppress_callback_exceptions=True, prevent_initial_callbacks=True)
server = app.server

app.layout = html.Div(children = [dcc.ConfirmDialog(id='confirm', message='DATA SAVED!!'),
            html.Center(html.Div(id='output-confirm')),
            html.Center(html.Div(children = [dcc.Input(id='input-on-submit', type='text', placeholder="Enter a ticker", style={'width': '20%'}), 
            dcc.Input(id='input-on-submit1', type='number', placeholder='Number of OHLCs (MAX1000)', max=1000, style={'width': '20%'}),
            dcc.Dropdown(id='demo-dropdown', options=[{'label': '1 Minute', 'value': '1Min'},
            {'label': '5 Minutes', 'value': '5Min'},
            {'label': '15 Minutes', 'value': '15Min'},
            {'label': '1 Day', 'value': 'day'}], placeholder="Span of each OHLC", value = "day", style={'width': '40%'})])),
            html.Br(),
            html.Div(html.Center(html.Button('Submit', id='submit-val', n_clicks=0))),
            html.Br(),
            html.Div(id='nouse', style={'display':'none'}),      # dummy boxes

            html.Div(dcc.Graph(id="graph")),
            html.Div(dcc.Graph(id='carGraph')), #Graph that displays all data
            html.Div(dcc.Graph(id='filterGraph')), #Graph that shows only filtered data
            html.Div(id='display'),  #To show format of selectData
            html.Br(),
            html.Div(html.Center(children = [html.Button("Grab Data", id='submit-val1', n_clicks=0)])),
            html.Div(html.Hr()),
            html.Div(html.Center(html.P(html.B("Short Time Fourier Transform")))),
            html.Div(html.Center(children=[dcc.Input(id='fs', type='number', 
                                                placeholder='Sampling frequency', max=1000, style={'width': '25%'}),
                                            dcc.Input(id='fft', type='number', 
                                                placeholder='Number of samples**', max=1000, style={'width': '25%', "margin":"10px"}),
                                            html.Button("Plot STFT", id='submit-val12', n_clicks=0, style={"margin":"20px"})])),
            html.Div(html.Center(html.P("**Number of samples in the fast fourier transform. Setting that value is a tradeoff between the time resolution and frequency resolution you want."))),
            html.Div(html.Hr()),
            html.Center(html.Div(id='textarea-example-output', style={'whiteSpace': 'pre-line'})),
            html.Div(id='nouse1', style={'display':'none'}),  # dummy boxes
            html.Div(id='nouse12', style={'display':'none'}
                        )  # dummy boxes

  
]) 

@app.callback(
        Output("graph", "figure"),
        Input('submit-val', 'n_clicks'),
        Input('input-on-submit1', "value"),
        Input('demo-dropdown', 'value'), 
        State('input-on-submit', "value")
    )
def display_candlestick(n_clicks, tspan, itspan, ticker):

    if n_clicks>0 and ticker!= None :
        api = alpaca.REST('PK7Z5SUF67ICPDK04R2M', 'ITAqIWxumbD67keejeh7yXTnrgSfnlZZZiXb759t', 'https://paper-api.alpaca.markets')
        df = api.get_barset(ticker, itspan, limit=tspan).df[ticker]

        s = df.index.to_pydatetime()
    
        fig = make_subplots(specs=[[{"secondary_y": True}]])

        # include candlestick with rangeselector
        fig.add_trace(go.Candlestick(x=s,
            open=df['open'], high=df['high'],
            low=df['low'], close=df['close']), secondary_y=False)

        # include a go.Bar trace for volumes
        fig.add_trace(go.Bar(x=s, y=df['volume'], marker_color='rgb(158,202,225)', marker_line_color='rgb(8,48,107)',
            marker_line_width=1.5, opacity=0.4),
            secondary_y=True)

        fig.layout.yaxis2.showgrid=False
        fig.update_yaxes(title_text="<b>VOLUME</b>", secondary_y=True)
        fig.update_yaxes(title_text="<b>STOCK PRICE</b>", secondary_y=False)
        fig.update_yaxes(automargin=True)

        fig.update_layout(showlegend=False, title=ticker, xaxis = {"showspikes": True}, yaxis = {"showspikes": True}, height=600)
        fig.update_layout(xaxis=dict(rangeselector=dict(
            buttons=list([
                dict(count=1,
                     label="1m",
                     step="month",
                     stepmode="backward"),
                dict(count=6,
                     label="6m",
                     step="month",
                     stepmode="backward"),
                dict(count=1,
                     label="YTD",
                     step="year",
                     stepmode="todate"),
                dict(count=1,
                     label="1y",
                     step="year",
                     stepmode="backward"),
                dict(step="all")
            ])
        ),
        rangeslider=dict(visible=True), type="date"), paper_bgcolor="LightSteelBlue")
        return fig

    else:
        fig2 = make_subplots(specs=[[{"secondary_y": True}]])
        fig2.layout.yaxis2.showgrid=False
        #fig2.update_yaxes(title_text="<b>VOLUME</b>", secondary_y=True)
        fig2.update_yaxes(title_text="<b>STOCK PRICE</b>", secondary_y=False)
        fig2.update_layout(#xaxis_rangeslider_visible='slider' in togg,
                 showlegend=False, title="Ticker")            
        return fig2

@app.callback(Output('confirm', 'displayed'),
              Input('submit-val1', 'n_clicks'))
def display_confirm(n_clicks1):
    if n_clicks1>0:
        return True
    return False

@app.callback(Output('output-confirm', 'children'),
              Input('submit-val1', 'n_clicks'))
def update_output(n_clicks):
    if n_clicks:
        return 'You Grabbed the Data {} times in this session!'.format(n_clicks)

@app.callback(Output('nouse1', 'children'),
        Input("submit-val1", 'n_clicks'),
        State("nouse", 'children'),
        State('input-on-submit1', "value"),
        State('demo-dropdown', 'value'), 
        State('input-on-submit', "value")
)
def making_dataset(n_clicks1, pts, tspan, itspan, ticker):
    if n_clicks1:    
        api = alpaca.REST('PK7Z5SUF67ICPDK04R2M', 'ITAqIWxumbD67keejeh7yXTnrgSfnlZZZiXb759t', 'https://paper-api.alpaca.markets')
        df = api.get_barset(ticker, itspan, limit=tspan).df[ticker]

        pts = yaml.load(pts)
        rang1 = pts[0]['text']
        rang2 = pts[len(pts)-1]['text']
        rang = [rang1, rang2]

        df.index = [x.strftime('%Y-%m-%d %H:%M:%S') for x in df.index]
        i = 0
        for str_data_time in rang:
            output = str_data_time.split(".")[0]
            rang[i] = output
            i += 1
        df = df.truncate(before = rang1, after = rang2) 
        
        df3 = pd.DataFrame({'SYMBOL': ticker,
                   'TIME': df.index, 'OPEN': df['open'],
                   'HIGH': df['high'], 'LOW': df['low'],
                   'CLOSE': df['close'], 'VOLUME': df['volume']})
        writer = pd.ExcelWriter(r"X:\Upwork\projects\plotting_trade_data\do_not_open.xlsx", engine='openpyxl')
        # try to open an existing workbook
        writer.book = load_workbook(r'X:\Upwork\projects\plotting_trade_data\do_not_open.xlsx')
        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # read existing file
        reader = pd.read_excel(r'X:\Upwork\projects\plotting_trade_data\do_not_open.xlsx', engine='openpyxl')
        # write out the new sheet
        df3.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
        writer.close()
        n_clicks1 = 0

        filename = r"X:\Upwork\projects\plotting_trade_data\do_not_open.xlsx"
        wb1 = xl.load_workbook(filename) 
        ws1 = wb1.worksheets[0] 
  
        # opening the destination excel file  
        filename1 = r"X:\Upwork\projects\plotting_trade_data\data_ohlc.xlsx"
        wb2 = xl.load_workbook(filename1) 
        ws2 = wb2.active 
  
        # calculate total number of rows and  
        # columns in source excel file 
        mr = ws1.max_row 
        mc = ws1.max_column 
  
        # copying the cell values from source  
        # excel file to destination excel file 
        for i in range (1, mr + 1): 
            for j in range (1, mc + 1): 
                # reading cell value from source excel file 
                c = ws1.cell(row = i, column = j) 
  
                # writing the read value to destination excel file 
                ws2.cell(row = i, column = j).value = c.value 
  
        # saving the destination excel file 
        wb2.save(str(filename1)) 

        return html.Div(html.H4(children="Running!!"))
 

@app.callback(
    Output('textarea-example-output', 'children'),
    Input('submit-val1', 'n_clicks'),
    Input('submit-val', 'n_clicks')
    )
def update_tables(n_clicks1, n_clicks):
    if n_clicks>0 and n_clicks1>=0:    
        df11 = pd.read_excel(r'X:\Upwork\projects\plotting_trade_data\do_not_open.xlsx',engine='openpyxl')  # pip3 install xlrd
        return html.Div(children=[
            html.Br(),
            html.Br(),
            html.H4(children='Last 10 saved records (Press submit to refresh)'),
                html.Table([
            html.Thead(
                html.Tr([html.Th(col) for col in df11.columns])
            ),
            html.Tbody([
                html.Tr([
                    html.Td(df11.iloc[i][col]) for col in df11.columns
                    ]) for i in range(len(df11)-10, len(df11))
                ])
            ])
        ]) 


@app.callback(Output('carGraph','figure'), Input('submit-val','n_clicks'),
            State('input-on-submit1', "value"),
            State('demo-dropdown', 'value'), 
            State('input-on-submit', "value"))            
def testfunc(clicks, tspan, itspan, ticker):
    api = alpaca.REST('PK7Z5SUF67ICPDK04R2M', 'ITAqIWxumbD67keejeh7yXTnrgSfnlZZZiXb759t', 'https://paper-api.alpaca.markets')
    df = api.get_barset(ticker, itspan, limit=tspan).df[ticker]    
    k = df.index.to_pydatetime()
    
    trace1 = go.Scatter(x=[k[i].date() for i in range(len(k))], y=df['close'],mode='markers+lines',text=[x.strftime('%Y-%m-%d %H:%M:%S') for x in df.index])
    layout = go.Layout(title='Use lasso or box tool to select', xaxis=dict(rangeslider=dict(visible=True), type="category", showgrid=False), yaxis=dict(title_text="<b>CLOSING PRICE</b>", showgrid=False), paper_bgcolor="LightSteelBlue", height = 600)

    return {'data':[trace1],'layout':layout}

# Show result of selecting data with either box select or lasso
@app.callback(Output('display','children'), Output('nouse','children'), [Input('carGraph','selectedData')])
def selectData(selectData):
    if selectData:
        return str('Selected Points: {}'.format(selectData)), str(selectData['points'])
    else:
        return str('None selected')

#Extract the 'text' component and use it to filter the dataframe and then create another graph
@app.callback(Output('filterGraph','figure'), [Input('carGraph','selectedData')], State('input-on-submit1', "value"),
            State('demo-dropdown', 'value'), State('input-on-submit', "value"))
def selectData3(selectData, tspan, itspan, ticker):
    if selectData:
        api = alpaca.REST('PK7Z5SUF67ICPDK04R2M', 'ITAqIWxumbD67keejeh7yXTnrgSfnlZZZiXb759t', 'https://paper-api.alpaca.markets')
        df4 = api.get_barset(ticker, itspan, limit=tspan).df[ticker]
        filtList = []
        for i in range(len(selectData['points'])):
            filtList.append(selectData['points'][i]['text'])
        df4['time'] = [x.strftime('%Y-%m-%d %H:%M:%S') for x in df4.index]
        filtList = [str(x) for x in filtList]
        df4['time']=[str(x) for x in df4['time']]

        filtCars = df4[df4['time'].isin(filtList)]
    
        trace2 = go.Scatter(x=filtCars.index.to_pydatetime(), y=filtCars['close'], mode='markers+lines', text=filtCars.index)
        layout2 = go.Layout(title='Grabbed Data')

        return {'data':[trace2],'layout':layout2}

    else:
        trace2 = go.Scatter()
        layout2 = go.Layout()
        return {'data':[trace2],'layout':layout2} 

# Analysis plots
@app.callback(Output('nouse12','children'), 
            Input('submit-val12','n_clicks'), 
            State("nouse", 'children'), 
            State('input-on-submit1', "value"),
            State('demo-dropdown', 'value'), 
            State('input-on-submit', "value"),
            State('fs', 'value'), 
            State('fft', "value")
            )
def graphs_analysis(n_clicks2, pts1, tspan, itspan, ticker, fs, fft_size):
    if n_clicks2>0:
        api = alpaca.REST('PK7Z5SUF67ICPDK04R2M', 'ITAqIWxumbD67keejeh7yXTnrgSfnlZZZiXb759t', 'https://paper-api.alpaca.markets')
        df = api.get_barset(ticker, itspan, limit=tspan).df[ticker]
        pts1 = yaml.load(pts1)
        rang1 = pts1[0]['text']
        rang2 = pts1[len(pts1)-1]['text']

        df = df.truncate(before = rang1, after = rang2) 

        data = df['close'].to_numpy()   # a numpy array containing the signal to be processed
        #fs = 800
        #fft_size = 500
        overlap_fac = 0.5
        hop_size = np.int32(np.floor(fft_size * (1-overlap_fac)))   
        pad_end_size = fft_size          # the last segment can overlap the end of the data array by no more than one window size
        total_segments = np.int32(np.ceil(len(data) / np.float32(hop_size)))    
        t_max = len(data) / np.float32(fs)
 
        window = np.hanning(fft_size)  # our half cosine window
        inner_pad = np.zeros(fft_size) # the zeros which will be used to double each segment size
 
        proc = np.concatenate((data, np.zeros(pad_end_size)))              # the data to process
        result = np.empty((total_segments, fft_size), dtype=np.float32)    # space to hold the result
 
        for i in range(total_segments):                      # for each segment
            current_hop = hop_size * i                        # figure out the current segment offset
            segment = proc[current_hop:current_hop+fft_size]  # get the current segment
            windowed = segment * window                       # multiply by the half cosine function
            padded = np.append(windowed, inner_pad)           # add 0s to double the length of the data
            spectrum = np.fft.fft(padded) / fft_size          # take the Fourier Transform and scale by the number of samples
            autopower = np.abs(spectrum * np.conj(spectrum))  # find the autopower spectrum
            result[i, :] = autopower[:fft_size]               # append to the results array
 
        result = 20*np.log10(result)          # scale to db
        result = np.clip(result, -40, 200)    # clip values
        
        root = tkinter.Tk()
        root.wm_title("STFT")

        fig = Figure(figsize=(5, 4), dpi=100)

        fig.add_subplot(111).imshow(result, origin='lower', cmap='jet', interpolation='nearest', aspect='auto')

        canvas = FigureCanvasTkAgg(fig, master=root)  # A tk.DrawingArea.
        canvas.draw()
        canvas.get_tk_widget().pack(side=tkinter.TOP, fill=tkinter.BOTH, expand=1)

        toolbar = NavigationToolbar2Tk(canvas, root)
        toolbar.update()
        canvas.get_tk_widget().pack(side=tkinter.TOP, fill=tkinter.BOTH, expand=1)
        
        def _quit():
            root.quit()     # stops mainloop
            root.destroy()  # this is necessary on Windows to prevent
                            # Fatal Python Error: PyEval_RestoreThread: NULL tstate

        button = tkinter.Button(master=root, text="QUIT", command=_quit)
        button.pack(side=tkinter.BOTTOM)
        tkinter.mainloop()
        return str("Done!")
    else:
        return str("Failed")

if __name__ == '__main__':
    app.run_server(threaded=True)
